# # encoding: utf-8
# # @author: MrZhou
# # @file: excel_handler.py
# # @time: 2022/9/28 16:19
# # @desc:
# """excel操作"""
import openpyxl
import data
import os


class ExcelHandler:
    def __init__(self, fpath):
        self.fpath = fpath

    def read(self, sheet_name):
        """读取数据"""
        # 打开文件
        wb = openpyxl.load_workbook(self.fpath)
        # 获取表格
        ws = wb[sheet_name]
        data = list(ws.values)
        # 关闭文件
        wb.close()
        header = data[0]
        all_data = []
        for row in data[1:]:
            row_dict = dict(zip(header, row))
            all_data.append(row_dict)
        return all_data

    def write(self, sheet_name, data, row, column):
        """写入excel数据"""
        wb = openpyxl.load_workbook(self.fpath)
        # 获取表格
        ws = wb[sheet_name]
        ws.cell(row=row, column=column).value = data
        # 通过workbook 保存和关闭
        wb.save(self.fpath)
        wb.close()


# if __name__ == '__main__':
#     ex = 'E:\KT03\jiekou\data\APICase.xlsx'
#     xls = ExcelHandler(ex)
#     excel_data = xls.read('register')
#     # excel_data
#     print(excel_data)
